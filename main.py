# import os
# import time
# import datetime
# from datetime import date, timedelta
# import re
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import (
#     TimeoutException, NoSuchElementException, StaleElementReferenceException
# )
# from selenium.webdriver.chrome.options import Options
# import gspread
# from google.oauth2.service_account import Credentials
# import json
# from typing import List, Dict, Optional, Tuple

# class MontgomeryCountyScraper:
#     def __init__(self, headless=True):
#         self.headless = headless
#         self.driver = None
#         self.setup_driver()
        
#     def setup_driver(self):
#         """Initialize the Chrome driver with appropriate options"""
#         chrome_options = Options()
#         if self.headless:
#             chrome_options.add_argument("--headless")
#         chrome_options.add_argument("--no-sandbox")
#         chrome_options.add_argument("--disable-dev-shm-usage")
#         chrome_options.add_argument("--disable-gpu")
#         chrome_options.add_argument("--window-size=1920,1080")
#         chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        
#         self.driver = webdriver.Chrome(options=chrome_options)
#         self.driver.implicitly_wait(2)
        
#     def retry_action(self, action, max_attempts=3, delay=2):
#         """Retry a specific action with delays between attempts"""
#         for attempt in range(max_attempts):
#             try:
#                 return action()
#             except (TimeoutException, StaleElementReferenceException) as e:
#                 if attempt == max_attempts - 1:
#                     raise e
#                 time.sleep(delay)
                
#     def wait_for_element(self, by, value, timeout=10):
#         """Wait for an element to be present"""
#         return WebDriverWait(self.driver, timeout).until(
#             EC.presence_of_element_located((by, value))
#         )
        
#     def wait_for_element_clickable(self, by, value, timeout=10):
#         """Wait for an element to be clickable"""
#         return WebDriverWait(self.driver, timeout).until(
#             EC.element_to_be_clickable((by, value))
#         )
        
#     def navigate_to_search_page(self):
#         """Navigate to the case search page"""
#         self.driver.get("https://courtsapp.montcopa.org/psi3/v/search/case")
#         time.sleep(3)  # Initial page load
        
#     def perform_search(self, start_date="01/01/2025", end_date=None):
#         """Perform the search with the given date range"""
#         if end_date is None:
#             end_date = date.today().strftime("%m/%d/%Y")
            
#         try:
#             # Wait for search box and enter search term
#             search_box = self.wait_for_element(By.ID, "Q")
#             search_box.clear()
#             search_box.send_keys("probate")
#             time.sleep(2)
            
#             # Set filing date from
#             from_date = self.wait_for_element(By.ID, "FilingDateFrom")
#             from_date.clear()
#             from_date.send_keys(start_date)
#             time.sleep(2)
            
#             # Set filing date to
#             to_date = self.wait_for_element(By.ID, "FilingDateTo")
#             to_date.clear()
#             to_date.send_keys(end_date)
#             time.sleep(2)
            
#             # Click search button
#             search_button = self.wait_for_element_clickable(
#                 By.XPATH, "//button[contains(@class, 'fa-search') and contains(text(), 'Search')]"
#             )
#             search_button.click()
#             time.sleep(3)
            
#             # Wait for results table
#             self.wait_for_element(By.ID, "gridViewResults")
#             return True
            
#         except Exception as e:
#             print(f"Error performing search: {e}")
#             return False
            
#     def extract_case_details(self, case_url):
#         """Extract case details from the detail page"""
#         self.driver.get(case_url)
#         time.sleep(3)
        
#         case_data = {
#             "case_number": "",
#             "last_filing_date": "",
#             "personal_representatives": [],
#             "case_foundation_parties_address": ""
#         }
        
#         try:
#             # Extract case number
#             try:
#                 case_number_element = self.wait_for_element(
#                     By.XPATH, "//*[@id='page-content-wrapper']/div/div[1]/div[1]/div[1]/table/tbody/tr[1]/td"
#                 )
#                 case_data["case_number"] = case_number_element.text.strip()
#             except:
#                 pass
                
#             # Extract last filing date
#             try:
#                 last_filing_element = self.wait_for_element(
#                     By.XPATH, "//*[@id='page-content-wrapper']/div/div[1]/div[1]/div[1]/table/tbody/tr[3]/td"
#                 )
#                 case_data["last_filing_date"] = last_filing_element.text.strip()
#             except:
#                 pass
                
#             # Extract personal representatives
#             try:
#                 reps_table = self.wait_for_element(By.ID, "table_PersonalRepresentatives")
#                 rows = reps_table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip header
                
#                 for row in rows:
#                     cells = row.find_elements(By.TAG_NAME, "td")
#                     if len(cells) >= 4:  # Ensure we have enough cells
#                         rep = {
#                             "name": cells[1].text.strip() if len(cells) > 1 else "",
#                             "role": cells[2].text.strip() if len(cells) > 2 else "",
#                             "address": cells[3].text.strip().replace("\n", ", ") if len(cells) > 3 else ""
#                         }
#                         case_data["personal_representatives"].append(rep)
#             except:
#                 pass
                
#             # Extract case foundation parties address
#             try:
#                 parties_table = self.wait_for_element(By.ID, "table_CaseFoundationParties")
#                 rows = parties_table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip header
                
#                 if rows:
#                     first_party_cells = rows[0].find_elements(By.TAG_NAME, "td")
#                     if len(first_party_cells) >= 5:  # Address is in the 5th column
#                         case_data["case_foundation_parties_address"] = first_party_cells[4].text.strip().replace("\n", ", ")
#             except:
#                 pass
                
#         except Exception as e:
#             print(f"Error extracting case details: {e}")
            
#         return case_data
        
#     def get_search_results(self):
#         """Get all search results from the current page"""
#         results = []
#         try:
#             table = self.wait_for_element(By.ID, "gridViewResults")
#             rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip header if exists
            
#             for row in rows:
#                 try:
#                     select_link = row.find_element(By.TAG_NAME, "a")
#                     case_url = select_link.get_attribute("href")
#                     results.append(case_url)
#                 except:
#                     continue
                    
#         except Exception as e:
#             print(f"Error getting search results: {e}")
            
#         return results
        
#     def has_next_page(self):
#         """Check if there's a next page of results"""
#         try:
#             next_button = self.driver.find_element(
#                 By.XPATH, "//a[contains(@href, 'Skip=') and contains(text(), 'Next')]"
#             )
#             return True
#         except:
#             return False
            
#     def go_to_next_page(self):
#         """Navigate to the next page of results"""
#         try:
#             next_button = self.wait_for_element_clickable(
#                 By.XPATH, "//a[contains(@href, 'Skip=') and contains(text(), 'Next')]"
#             )
#             next_button.click()
#             time.sleep(3)
#             self.wait_for_element(By.ID, "gridViewResults")
#             return True
#         except Exception as e:
#             print(f"Error going to next page: {e}")
#             return False
            
#     def scrape_cases(self, start_date="01/01/2025", end_date=None):
#         """Main method to scrape all cases"""
#         if end_date is None:
#             end_date = date.today().strftime("%m/%d/%Y")
            
#         self.navigate_to_search_page()
        
#         if not self.perform_search(start_date, end_date):
#             print("Search failed")
#             return []
            
#         all_case_data = []
#         page = 1
        
#         while True:
#             print(f"Scraping page {page}")
#             case_urls = self.get_search_results()
            
#             for case_url in case_urls:
#                 case_data = self.extract_case_details(case_url)
                
#                 # Only include cases with personal representatives
#                 if case_data["personal_representatives"]:
#                     all_case_data.append(case_data)
                
#                 # Go back to search results
#                 self.driver.back()
#                 time.sleep(3)
#                 self.wait_for_element(By.ID, "gridViewResults")
                
#             # Check for next page
#             if not self.has_next_page():
#                 break
                
#             if not self.go_to_next_page():
#                 break
                
#             page += 1
            
#         return all_case_data
        
#     def close(self):
#         """Close the browser"""
#         if self.driver:
#             self.driver.quit()


# class GoogleSheetsHandler:
#     def __init__(self, credentials_json, spreadsheet_id):
#         self.credentials_json = credentials_json
#         self.spreadsheet_id = spreadsheet_id
#         self.client = None
#         self.spreadsheet = None
#         self.setup_client()
        
#     def setup_client(self):
#         """Set up the Google Sheets client"""
#         try:
#             scopes = [
#                 "https://www.googleapis.com/auth/spreadsheets",
#                 "https://www.googleapis.com/auth/drive"
#             ]
#             credentials = Credentials.from_service_account_info(
#                 json.loads(self.credentials_json), scopes=scopes
#             )
#             self.client = gspread.authorize(credentials)
#             self.spreadsheet = self.client.open_by_key(self.spreadsheet_id)
#         except Exception as e:
#             print(f"Error setting up Google Sheets client: {e}")
#             raise
            
#     def get_or_create_sheet(self, sheet_name):
#         """Get or create a sheet with the given name"""
#         try:
#             worksheet = self.spreadsheet.worksheet(sheet_name)
#             print(f"Found existing sheet: {sheet_name}")
#         except gspread.exceptions.WorksheetNotFound:
#             print(f"Creating new sheet: {sheet_name}")
#             worksheet = self.spreadsheet.add_worksheet(
#                 title=sheet_name, rows=1000, cols=20
#             )
            
#             # Set up headers
#             headers = [
#                 "Case Number", 
#                 "Last Filing Date", 
#                 "Representative Name", 
#                 "Role", 
#                 "Address", 
#                 "Case Foundation Parties Address"
#             ]
#             worksheet.update("A1:F1", [headers])
            
#             # Format headers
#             worksheet.format("A1:F1", {
#                 "textFormat": {"bold": True},
#                 "backgroundColor": {"red": 0.9, "green": 0.9, "blue": 0.9}
#             })
            
#             # Freeze header row
#             worksheet.freeze(rows=1)
            
#         return worksheet
        
#     def update_sheet(self, sheet_name, case_data):
#         """Update the sheet with the given case data"""
#         worksheet = self.get_or_create_sheet(sheet_name)
        
#         # Prepare data for writing
#         rows = []
#         for case in case_data:
#             for rep in case["personal_representatives"]:
#                 row = [
#                     case["case_number"],
#                     case["last_filing_date"],
#                     rep["name"],
#                     rep["role"],
#                     rep["address"],
#                     case["case_foundation_parties_address"]
#                 ]
#                 rows.append(row)
                
#         if rows:
#             # Clear existing data (except headers)
#             worksheet.batch_clear(["A2:F1000"])
            
#             # Write new data
#             worksheet.update(f"A2:F{len(rows)+1}", rows)
            
#             # Auto-resize columns
#             worksheet.columns_auto_resize(0, 5)  # Columns A to F
            
#         print(f"Updated sheet {sheet_name} with {len(rows)} rows")
        
#     def get_last_scraped_date(self, sheet_name):
#         """Get the last date that was scraped from a sheet"""
#         try:
#             worksheet = self.spreadsheet.worksheet(sheet_name)
#             dates = worksheet.col_values(2)  # Column B has filing dates
#             if len(dates) > 1:  # Skip header
#                 # Return the latest date
#                 return max(
#                     [self.parse_date(date_str) for date_str in dates[1:] if date_str],
#                     default=None
#                 )
#         except:
#             pass
#         return None
        
#     def parse_date(self, date_str):
#         """Parse a date string into a date object"""
#         try:
#             # Handle various date formats
#             for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
#                 try:
#                     return datetime.datetime.strptime(date_str, fmt).date()
#                 except ValueError:
#                     continue
#         except:
#             pass
#         return None


# def main():
#     # Configuration
#     HEADLESS = True
#     START_DATE = "01/01/2025"
    
#     # Get credentials from environment
#     google_credentials = os.environ.get("GOOGLE_CREDENTIALS")
#     spreadsheet_id = os.environ.get("SPREADSHEET_ID")
    
#     if not google_credentials or not spreadsheet_id:
#         print("Error: GOOGLE_CREDENTIALS and SPREADSHEET_ID environment variables are required")
#         return
        
#     # Initialize Google Sheets handler
#     sheets_handler = GoogleSheetsHandler(google_credentials, spreadsheet_id)
    
#     # Determine the current month for the sheet name
#     current_month = date.today().strftime("%Y-%m")
    
#     # Check if we have a previous scrape for this month
#     last_date = sheets_handler.get_last_scraped_date(current_month)
    
#     # If we have a previous scrape, start from the day after the last date
#     if last_date:
#         start_date = (last_date + timedelta(days=1)).strftime("%m/%d/%Y")
#     else:
#         start_date = START_DATE
        
#     # Initialize scraper
#     scraper = MontgomeryCountyScraper(headless=HEADLESS)
    
#     try:
#         # Scrape cases
#         case_data = scraper.scrape_cases(start_date=start_date)
        
#         if case_data:
#             # Update Google Sheet
#             sheets_handler.update_sheet(current_month, case_data)
#             print(f"Successfully processed {len(case_data)} cases")
#         else:
#             print("No cases found or no new cases since last scrape")
            
#     except Exception as e:
#         print(f"Error during scraping: {e}")
        
#     finally:
#         scraper.close()


# if __name__ == "__main__":
#     main()

import os
import time
import datetime
from datetime import date, timedelta
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2 import service_account
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

# -----------------------------
# Scraper Definition (same as your code, trimmed sleeps for CI speed)
# -----------------------------
class MontgomeryCountyScraper:
    def __init__(self, headless=True):
        self.headless = headless
        self.driver = None
        self.setup_driver()
    def setup_driver(self):
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")

        if os.environ.get("SELENIUM_REMOTE_URL"):  # GitHub Actions
            remote_url = os.environ["SELENIUM_REMOTE_URL"]
            self.driver = webdriver.Remote(
                command_executor=remote_url,
                options=chrome_options
            )
        else:  # local dev
            if self.headless:
                chrome_options.add_argument("--headless=new")
            self.driver = webdriver.Chrome(options=chrome_options)

        self.driver.implicitly_wait(2)

    def wait(self, by, value, timeout=10):
        return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, value)))
    def navigate(self):
        self.driver.get("https://courtsapp.montcopa.org/psi3/v/search/case")
        time.sleep(2)
    def perform_search(self, start_date, end_date):
        try:
            self.wait(By.ID,"Q").send_keys("probate")
            fdate = self.wait(By.ID,"FilingDateFrom"); fdate.clear(); fdate.send_keys(start_date)
            tdate = self.wait(By.ID,"FilingDateTo"); tdate.clear(); tdate.send_keys(end_date)
            btn = self.driver.find_element(By.XPATH,"//button[contains(text(),'Search')]")
            btn.click(); time.sleep(2)
            self.wait(By.ID,"gridViewResults")
            return True
        except Exception as e:
            print("Search error:", e); return False
    def get_results(self):
        urls=[]
        table=self.driver.find_element(By.ID,"gridViewResults")
        rows=table.find_elements(By.TAG_NAME,"tr")[1:]
        for r in rows:
            try: urls.append(r.find_element(By.TAG_NAME,"a").get_attribute("href"))
            except: continue
        return urls
    def extract(self, url):
        self.driver.get(url); time.sleep(1)
        case={"case_number":"","last_filing_date":"","personal_representatives":[],"case_foundation_parties_address":""}
        try:
            case["case_number"]=self.driver.find_element(By.XPATH,"//table/tbody/tr[1]/td").text.strip()
        except: pass
        try:
            case["last_filing_date"]=self.driver.find_element(By.XPATH,"//table/tbody/tr[3]/td").text.strip()
        except: pass
        try:
            reps=self.driver.find_element(By.ID,"table_PersonalRepresentatives").find_elements(By.TAG_NAME,"tr")[1:]
            for row in reps:
                cells=row.find_elements(By.TAG_NAME,"td")
                if len(cells)>=4:
                    case["personal_representatives"].append({
                        "name":cells[1].text.strip(),
                        "role":cells[2].text.strip(),
                        "address":cells[3].text.strip().replace("\n",", ")
                    })
        except: pass
        try:
            rows=self.driver.find_element(By.ID,"table_CaseFoundationParties").find_elements(By.TAG_NAME,"tr")[1:]
            if rows:
                cells=rows[0].find_elements(By.TAG_NAME,"td")
                if len(cells)>=5:
                    case["case_foundation_parties_address"]=cells[4].text.strip().replace("\n",", ")
        except: pass
        return case
    def has_next(self):
        try: self.driver.find_element(By.XPATH,"//a[contains(text(),'Next')]"); return True
        except: return False
    def next_page(self):
        try:
            self.driver.find_element(By.XPATH,"//a[contains(text(),'Next')]").click()
            time.sleep(1); self.wait(By.ID,"gridViewResults"); return True
        except: return False
    def scrape(self,start_date,end_date):
        self.navigate()
        if not self.perform_search(start_date,end_date): return []
        data=[]; page=1
        while True:
            for url in self.get_results():
                case=self.extract(url)
                if case["personal_representatives"]:
                    data.append(case)
                self.driver.back(); self.wait(By.ID,"gridViewResults")
            if not self.has_next() or not self.next_page(): break
            page+=1
        return data
    def close(self):
        if self.driver: self.driver.quit()

# -----------------------------
# Google Sheets Helpers
# -----------------------------
SCOPES=["https://www.googleapis.com/auth/spreadsheets"]

def load_service_account_info():
    raw=os.environ.get("GOOGLE_CREDENTIALS")
    if not raw: raise RuntimeError("Missing GOOGLE_CREDENTIALS")
    return json.loads(raw)

def sheets_client():
    info=load_service_account_info()
    creds=service_account.Credentials.from_service_account_info(info,scopes=SCOPES)
    return build("sheets","v4",credentials=creds).spreadsheets()

def get_last_scraped_date(svc, spreadsheet_id, sheet_name):
    try:
        # Column 2 = "Last Filing Date", start from row 2 (skip header)
        res = svc.values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!B2:B"
        ).execute()
        vals = [r[0] for r in res.get("values", []) if r]

        if not vals:
            return None

        parsed = []
        for d in vals:
            for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                try:
                    parsed.append(datetime.datetime.strptime(d, fmt).date())
                    break
                except ValueError:
                    continue
        return max(parsed) if parsed else None

    except Exception as e:
        print(f"⚠ Error reading last date from {sheet_name}: {e}")
        return None

def append_rows(svc, spreadsheet_id, sheet_name, rows):
    if not rows: return
    svc.values().append(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!A:Z",
        body={"values":rows},
        valueInputOption="USER_ENTERED"
    ).execute()
    print(f"✓ Appended {len(rows)} rows to {sheet_name}")

def normalize_date(datestr):
    """Convert various formats into YYYY-MM-DD"""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(datestr, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return datestr  # fallback unchanged


def ensure_sheet_exists(svc, spreadsheet_id, sheet_name):
    """Create a sheet if missing"""
    try:
        svc.get(spreadsheetId=spreadsheet_id).execute()
        sheets_metadata = svc.get(spreadsheetId=spreadsheet_id).execute()
        titles = [s["properties"]["title"] for s in sheets_metadata["sheets"]]
        if sheet_name not in titles:
            body = {"requests": [{"addSheet": {"properties": {"title": sheet_name}}}]}
            svc.batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
            print(f"✓ Created new sheet: {sheet_name}")
    except Exception as e:
        print(f"⚠ Could not ensure sheet {sheet_name}: {e}")


def update_summary(svc, spreadsheet_id, month_name, new_count):
    """Update monthly total instead of appending daily"""
    try:
        res = svc.values().get(
            spreadsheetId=spreadsheet_id,
            range="'Summary'!A:B"
        ).execute()
        rows = res.get("values", [])

        updated = False
        for i, row in enumerate(rows, start=1):
            if row and row[0] == month_name:
                # Update existing month count
                old_count = int(row[1]) if len(row) > 1 and row[1].isdigit() else 0
                new_total = old_count + new_count
                svc.values().update(
                    spreadsheetId=spreadsheet_id,
                    range=f"'Summary'!B{i}",
                    valueInputOption="USER_ENTERED",
                    body={"values": [[new_total]]}
                ).execute()
                updated = True
                print(f"✓ Updated Summary: {month_name} = {new_total}")
                break

        if not updated:
            # Append new month row
            svc.values().append(
                spreadsheetId=spreadsheet_id,
                range="'Summary'!A:B",
                valueInputOption="USER_ENTERED",
                body={"values": [[month_name, new_count]]}
            ).execute()
            print(f"✓ Added Summary row: {month_name} = {new_count}")

    except Exception as e:
        print(f"⚠ Error updating Summary: {e}")


# -----------------------------
# Main Logic
# -----------------------------
def main():
    spreadsheet_id = os.environ.get("SPREADSHEET_ID")
    svc = sheets_client()
    today = date.today()

    # Monthly sheet in "Sep 2025" format
    month_sheet = today.strftime("%b %Y")
    ensure_sheet_exists(svc, spreadsheet_id, month_sheet)
    ensure_sheet_exists(svc, spreadsheet_id, "All Data")
    ensure_sheet_exists(svc, spreadsheet_id, "Summary")


    last_date = get_last_scraped_date(svc, spreadsheet_id, month_sheet)

    if last_date:
        start_date = (last_date + timedelta(days=1)).strftime("%m/%d/%Y")
    else:
        # If no rows yet → start from 1st of month
        start_date = today.replace(day=1).strftime("%m/%d/%Y")

    end_date = today.strftime("%m/%d/%Y")

    print(f"Scraping {start_date} → {end_date} into {month_sheet}")

    scraper = MontgomeryCountyScraper(headless=True)
    try:
        data = scraper.scrape(start_date, end_date)
        rows = []
        for case in data:
            for rep in case["personal_representatives"]:
                rows.append([
                    case["case_number"],
                    normalize_date(case["last_filing_date"]),
                    rep["name"],
                    rep["role"],
                    rep["address"],
                    case["case_foundation_parties_address"],
                ])

        if rows:
            append_rows(svc, spreadsheet_id, month_sheet, rows)
            append_rows(svc, spreadsheet_id, "All Data", rows)
            update_summary(svc, spreadsheet_id, month_sheet, len(rows))
        else:
            print("✓ No new rows")
    finally:
        scraper.close()
if __name__=="__main__":
    main()
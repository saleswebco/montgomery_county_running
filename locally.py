import os
import time
import datetime
from datetime import date, timedelta
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException
)
from selenium.webdriver.chrome.options import Options
import json
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


class MontgomeryCountyScraper:
    def __init__(self, headless=True):
        self.headless = headless
        self.driver = None
        self.setup_driver()
        
    def setup_driver(self):
        """Initialize the Chrome driver with appropriate options"""
        chrome_options = Options()
        if self.headless:
            chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.implicitly_wait(2)
        
    def wait_for_element(self, by, value, timeout=10):
        """Wait for an element to be present"""
        return WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((by, value))
        )
        
    def wait_for_element_clickable(self, by, value, timeout=10):
        """Wait for an element to be clickable"""
        return WebDriverWait(self.driver, timeout).until(
            EC.element_to_be_clickable((by, value))
        )
        
    def navigate_to_search_page(self):
        """Navigate to the case search page"""
        self.driver.get("https://courtsapp.montcopa.org/psi3/v/search/case")
        time.sleep(3)  # Initial page load
        
    def perform_search(self, start_date="01/01/2025", end_date=None):
        """Perform the search with the given date range"""
        if end_date is None:
            end_date = date.today().strftime("%m/%d/%Y")
            
        try:
            search_box = self.wait_for_element(By.ID, "Q")
            search_box.clear()
            search_box.send_keys("probate")
            time.sleep(1)
            
            from_date = self.wait_for_element(By.ID, "FilingDateFrom")
            from_date.clear()
            from_date.send_keys(start_date)
            
            to_date = self.wait_for_element(By.ID, "FilingDateTo")
            to_date.clear()
            to_date.send_keys(end_date)
            
            search_button = self.wait_for_element_clickable(
                By.XPATH, "//button[contains(@class, 'fa-search') and contains(text(), 'Search')]"
            )
            search_button.click()
            time.sleep(3)
            
            self.wait_for_element(By.ID, "gridViewResults")
            return True
            
        except Exception as e:
            print(f"Error performing search: {e}")
            return False
            
    def extract_case_details(self, case_url):
        """Extract case details from the detail page"""
        self.driver.get(case_url)
        time.sleep(3)
        
        case_data = {
            "case_number": "",
            "last_filing_date": "",
            "personal_representatives": [],
            "case_foundation_parties_address": ""
        }
        
        try:
            try:
                case_number_element = self.wait_for_element(
                    By.XPATH, "//*[@id='page-content-wrapper']/div/div[1]/div[1]/div[1]/table/tbody/tr[1]/td"
                )
                case_data["case_number"] = case_number_element.text.strip()
            except:
                pass
                
            try:
                last_filing_element = self.wait_for_element(
                    By.XPATH, "//*[@id='page-content-wrapper']/div/div[1]/div[1]/div[1]/table/tbody/tr[3]/td"
                )
                case_data["last_filing_date"] = last_filing_element.text.strip()
            except:
                pass
                
            try:
                reps_table = self.wait_for_element(By.ID, "table_PersonalRepresentatives")
                rows = reps_table.find_elements(By.TAG_NAME, "tr")[1:]
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 4:
                        rep = {
                            "name": cells[1].text.strip(),
                            "role": cells[2].text.strip(),
                            "address": cells[3].text.strip().replace("\n", ", ")
                        }
                        case_data["personal_representatives"].append(rep)
            except:
                pass
                
            try:
                parties_table = self.wait_for_element(By.ID, "table_CaseFoundationParties")
                rows = parties_table.find_elements(By.TAG_NAME, "tr")[1:]
                if rows:
                    first_party_cells = rows[0].find_elements(By.TAG_NAME, "td")
                    if len(first_party_cells) >= 5:
                        case_data["case_foundation_parties_address"] = first_party_cells[4].text.strip().replace("\n", ", ")
            except:
                pass
                
        except Exception as e:
            print(f"Error extracting case details: {e}")
            
        return case_data
        
    def get_search_results(self):
        """Get all search results from the current page"""
        results = []
        try:
            table = self.wait_for_element(By.ID, "gridViewResults")
            rows = table.find_elements(By.TAG_NAME, "tr")[1:]
            for row in rows:
                try:
                    select_link = row.find_element(By.TAG_NAME, "a")
                    case_url = select_link.get_attribute("href")
                    results.append(case_url)
                except:
                    continue
                    
        except Exception as e:
            print(f"Error getting search results: {e}")
            
        return results
        
    def has_next_page(self):
        """Check if there's a next page of results"""
        try:
            self.driver.find_element(
                By.XPATH, "//a[contains(@href, 'Skip=') and contains(text(), 'Next')]"
            )
            return True
        except:
            return False
            
    def go_to_next_page(self):
        """Navigate to the next page of results"""
        try:
            next_button = self.wait_for_element_clickable(
                By.XPATH, "//a[contains(@href, 'Skip=') and contains(text(), 'Next')]"
            )
            next_button.click()
            time.sleep(3)
            self.wait_for_element(By.ID, "gridViewResults")
            return True
        except Exception as e:
            print(f"Error going to next page: {e}")
            return False
            
    def scrape_cases(self, start_date="01/01/2025", end_date=None):
        """Main method to scrape all cases"""
        if end_date is None:
            end_date = date.today().strftime("%m/%d/%Y")
            
        self.navigate_to_search_page()
        
        if not self.perform_search(start_date, end_date):
            print("Search failed")
            return []
            
        all_case_data = []
        page = 1
        
        while True:
            print(f"Scraping page {page}")
            case_urls = self.get_search_results()
            
            for case_url in case_urls:
                case_data = self.extract_case_details(case_url)
                if case_data["personal_representatives"]:
                    all_case_data.append(case_data)
                self.driver.back()
                time.sleep(2)
                self.wait_for_element(By.ID, "gridViewResults")
                
            if not self.has_next_page() or not self.go_to_next_page():
                break
            page += 1
            
        return all_case_data
        
    def close(self):
        if self.driver:
            self.driver.quit()


class ExcelHandler:
    def __init__(self, file_path="cases.xlsx"):
        self.file_path = file_path
        
    def get_or_create_sheet(self, wb, sheet_name):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            headers = [
                "Case Number", 
                "Last Filing Date", 
                "Representative Name", 
                "Role", 
                "Address", 
                "Case Foundation Parties Address"
            ]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)
                ws.cell(row=1, column=col).fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        return ws
        
    def get_last_scraped_date(self, sheet_name):
        if not os.path.exists(self.file_path):
            return None
        wb = load_workbook(self.file_path)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        dates = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=2).value]
        if not dates:
            return None
        try:
            parsed_dates = [self.parse_date(d) for d in dates if d]
            return max([d for d in parsed_dates if d], default=None)
        except:
            return None
        
    def parse_date(self, date_str):
        if isinstance(date_str, datetime.date):
            return date_str
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
            try:
                return datetime.datetime.strptime(str(date_str), fmt).date()
            except ValueError:
                continue
        return None
        
    def update_sheet(self, sheet_name, case_data):
        if os.path.exists(self.file_path):
            wb = load_workbook(self.file_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
                
        ws = self.get_or_create_sheet(wb, sheet_name)
        
        for case in case_data:
            for rep in case["personal_representatives"]:
                row = [
                    case["case_number"],
                    case["last_filing_date"],
                    rep["name"],
                    rep["role"],
                    rep["address"],
                    case["case_foundation_parties_address"]
                ]
                ws.append(row)
                
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
            
        wb.save(self.file_path)
        print(f"Updated {self.file_path} [{sheet_name}] with {len(case_data)} cases")


def main():
    HEADLESS = False  # 👈 Run with UI for checking
    START_DATE = "01/01/2025"
    FILE_PATH = "cases.xlsx"
    
    excel_handler = ExcelHandler(FILE_PATH)
    
    current_month = date.today().strftime("%Y-%m")
    last_date = excel_handler.get_last_scraped_date(current_month)
    
    if last_date:
        start_date = (last_date + timedelta(days=1)).strftime("%m/%d/%Y")
    else:
        start_date = START_DATE
        
    scraper = MontgomeryCountyScraper(headless=HEADLESS)
    
    try:
        case_data = scraper.scrape_cases(start_date=start_date)
        if case_data:
            excel_handler.update_sheet(current_month, case_data)
            print(f"Successfully processed {len(case_data)} cases")
        else:
            print("No cases found or no new cases since last scrape")
    except Exception as e:
        print(f"Error during scraping: {e}")
    finally:
        scraper.close()


if __name__ == "__main__":
    main()
